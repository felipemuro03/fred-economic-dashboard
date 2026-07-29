import re

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font


def _nome_aba_valido(nome: str) -> str:
    """Nomes de aba do Excel têm limite de 31 caracteres e não aceitam alguns símbolos."""
    limpo = re.sub(r"[\\/*?:\[\]]", "", nome)
    return limpo[:31]


def exportar_excel(series_selecionadas: dict, caminho_saida: str, rotulo_codigo: str = "Código"):
    """
    series_selecionadas: dict no formato
        {series_id: {"nome": str, "serie": pandas.Series, "nota": str}}
    Gera um .xlsx com uma aba "Resumo" (últimos valores) e uma aba por série
    (data/valor + gráfico de linha).
    """
    wb = Workbook()
    resumo = wb.active
    resumo.title = "Resumo"
    resumo.append(["Indicador", rotulo_codigo, "Última Data", "Último Valor", "Unidade", "Nota", "Link"])
    for celula in resumo[1]:
        celula.font = Font(bold=True)

    for series_id, dados in series_selecionadas.items():
        serie = dados["serie"].dropna()
        nome = dados.get("nome", series_id)
        nota = dados.get("nota", "")
        url = dados.get("url", "")
        unidade_badge = dados.get("unidade", {}).get("badge", "")

        if not serie.empty:
            ultima_data = serie.index[-1].strftime("%Y-%m-%d")
            ultimo_valor = float(serie.iloc[-1])
        else:
            ultima_data, ultimo_valor = "", None
        resumo.append(
            [nome, series_id, ultima_data, ultimo_valor, unidade_badge, nota, "Ver fonte" if url else ""]
        )
        if url:
            celula_link = resumo.cell(row=resumo.max_row, column=7)
            celula_link.hyperlink = url
            celula_link.font = Font(color="0563C1", underline="single")

        aba = wb.create_sheet(_nome_aba_valido(nome))
        aba.append(["Data", "Valor"])
        for data, valor in serie.items():
            aba.append([data.strftime("%Y-%m-%d"), float(valor)])

        grafico = LineChart()
        grafico.title = nome
        grafico.y_axis.title = "Valor"
        grafico.x_axis.title = "Data"
        n_linhas = len(serie) + 1
        valores_ref = Reference(aba, min_col=2, min_row=1, max_row=n_linhas)
        categorias_ref = Reference(aba, min_col=1, min_row=2, max_row=n_linhas)
        grafico.add_data(valores_ref, titles_from_data=True)
        grafico.set_categories(categorias_ref)
        aba.add_chart(grafico, "D2")

    wb.save(caminho_saida)
    return caminho_saida
